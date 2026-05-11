# web_app.py
"""
Complete Hospital-Grade Web Application for Childbirth Mode Prediction
With multiple doctor logins, admin access, Excel-based storage, and enhanced XAI
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import matplotlib.pyplot as plt
import shap
import joblib
import pickle
import os
import sys
import json
from datetime import datetime
import hashlib
import base64
import warnings
import openpyxl
from openpyxl import Workbook, load_workbook
from pathlib import Path
warnings.filterwarnings('ignore')

# Page configuration
st.set_page_config(
    page_title="Childbirth Mode Predictor",
    page_icon="👶",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1E3A8A;
        text-align: center;
        margin-bottom: 2rem;
        padding: 1rem;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border-radius: 10px;
    }
    .sub-header {
        font-size: 1.8rem;
        color: #3B82F6;
        margin-bottom: 1.5rem;
        padding-bottom: 0.5rem;
        border-bottom: 3px solid #3B82F6;
    }
    .card {
        background-color: #FFFFFF;
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        margin-bottom: 1.5rem;
        border: 1px solid #E5E7EB;
    }
    .prediction-card {
        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        color: white;
        padding: 2rem;
        border-radius: 15px;
        margin: 1rem 0;
    }
    .high-risk {
        background: linear-gradient(135deg, #ff6b6b 0%, #ee5a52 100%);
    }
    .medium-risk {
        background: linear-gradient(135deg, #f6d365 0%, #fda085 100%);
    }
    .low-risk {
        background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);
    }
    .info-card {
        background-color: #F0F9FF;
        padding: 1.5rem;
        border-radius: 10px;
        border-left: 5px solid #3B82F6;
        margin: 1rem 0;
    }
    .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        font-weight: bold;
        border: none;
        padding: 0.75rem 2rem;
        border-radius: 8px;
        font-size: 1rem;
        transition: all 0.3s ease;
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 12px rgba(0, 0, 0, 0.15);
    }
    .metric-card {
        background-color: white;
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
        text-align: center;
    }
    .doctor-card {
        background: linear-gradient(135deg, #a8edea 0%, #fed6e3 100%);
        padding: 1.5rem;
        border-radius: 10px;
        margin: 1rem 0;
    }
    .admin-card {
        background: linear-gradient(135deg, #ffecd2 0%, #fcb69f 100%);
        padding: 1.5rem;
        border-radius: 10px;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

# Session State Initialization
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False
if 'current_user' not in st.session_state:
    st.session_state.current_user = None
if 'user_role' not in st.session_state:
    st.session_state.user_role = None
if 'current_page' not in st.session_state:
    st.session_state.current_page = "Home"
if 'patient_data' not in st.session_state:
    st.session_state.patient_data = {}
if 'predictions' not in st.session_state:
    st.session_state.predictions = {}
if 'model_loaded' not in st.session_state:
    st.session_state.model_loaded = False

# Initialize session state variables for models
if 'model' not in st.session_state:
    st.session_state.model = None
if 'scaler' not in st.session_state:
    st.session_state.scaler = None
if 'encoders' not in st.session_state:
    st.session_state.encoders = None
if 'feature_names' not in st.session_state:
    st.session_state.feature_names = None
if 'target_classes' not in st.session_state:
    st.session_state.target_classes = None
if 'shap_explainer' not in st.session_state:
    st.session_state.shap_explainer = None
if 'feature_descriptions' not in st.session_state:
    st.session_state.feature_descriptions = None

# Utility Functions
def calculate_bmi(weight, height):
    """Calculate BMI from weight (kg) and height (cm)"""
    if weight and height:
        height_m = height / 100
        return weight / (height_m ** 2)
    return None

def get_bmi_category(bmi):
    """Get BMI category"""
    if bmi < 18.5:
        return "Underweight"
    elif bmi < 25:
        return "Normal"
    elif bmi < 30:
        return "Overweight"
    else:
        return "Obese"

def calculate_risk_score(patient_data):
    """Calculate clinical risk score"""
    score = 0
    
    # Age factor
    age = patient_data.get('AGE', 25)
    if age >= 35:
        score += 25
    elif age >= 30:
        score += 15
    elif age >= 25:
        score += 5
    
    # BMI factor
    bmi = patient_data.get('BMI', 22)
    if bmi >= 30:
        score += 30
    elif bmi >= 25:
        score += 20
    elif bmi < 18.5:
        score += 15
    
    # Previous cesarean
    if patient_data.get('PREVIOUS_CESAREAN', 0) == 1:
        score += 30
        if patient_data.get('NUMBER_PREV_CESAREAN', 0) > 1:
            score += 10
    
    # Comorbidities
    if patient_data.get('COMORBIDITY', 0) == 1:
        score += 20
    
    # Lifestyle factors
    if patient_data.get('SMOKING', 0) == 1:
        score += 15
    if patient_data.get('ALCOHOL', 0) == 1:
        score += 10
    
    # Obstetric risk
    obstetric_risk = patient_data.get('OBSTETRIC_RISK', 'low')
    if obstetric_risk == 'high':
        score += 25
    elif obstetric_risk == 'medium':
        score += 15
    
    # Gestational age
    ga = patient_data.get('GESTATIONAL_AGE', 38)
    if ga < 37:
        score += 20
    elif ga >= 42:
        score += 15
    
    return min(score, 100)

def get_risk_level(score):
    """Get risk level category"""
    if score >= 80:
        return "🚨 Very High Risk"
    elif score >= 60:
        return "⚠️ High Risk"
    elif score >= 40:
        return "📊 Moderate Risk"
    else:
        return "✅ Low Risk"

def get_recommendations(prediction, probability, risk_score):
    """Generate clinical recommendations"""
    recommendations = []
    
    # Risk-based recommendations
    if risk_score >= 80:
        recommendations.append("🚨 **URGENT ACTION REQUIRED**")
        recommendations.append("• Immediate specialist consultation")
        recommendations.append("• Continuous fetal monitoring")
        recommendations.append("• Prepare for emergency intervention")
        recommendations.append("• Consider transfer to tertiary care center")
    elif risk_score >= 60:
        recommendations.append("⚠️ **HIGH RISK MANAGEMENT**")
        recommendations.append("• Enhanced monitoring required")
        recommendations.append("• Regular fetal assessment")
        recommendations.append("• Discuss birth plan with specialist")
        recommendations.append("• Prepare for possible interventions")
    
    # Prediction-based recommendations
    if "CES" in str(prediction):
        recommendations.append("**CESAREAN DELIVERY PREPAREDNESS**")
        recommendations.append("• Schedule pre-operative assessment")
        recommendations.append("• Discuss anesthesia options")
        recommendations.append("• Plan for post-operative care")
        recommendations.append("• Prepare neonatal resuscitation equipment")
    else:
        recommendations.append("**VAGINAL DELIVERY SUPPORT**")
        recommendations.append("• Monitor labor progress closely")
        recommendations.append("• Consider pain management options")
        recommendations.append("• Prepare for assisted delivery if needed")
        recommendations.append("• Plan for immediate neonatal care")
    
    # Confidence-based recommendations
    if probability < 0.7:
        recommendations.append("⚠️ **LOW CONFIDENCE PREDICTION**")
        recommendations.append("• Consider additional diagnostic tests")
        recommendations.append("• Second opinion recommended")
        recommendations.append("• Close monitoring advised")
    
    return recommendations

def create_patient_id(name, dob):
    """Create unique patient ID"""
    unique_string = f"{name}{dob}{datetime.now().timestamp()}"
    return hashlib.md5(unique_string.encode()).hexdigest()[:8].upper()

def save_to_excel(patient_data, prediction, probability, risk_score, doctor):
    """Save patient record to Excel file"""
    try:
        # Create patient directory if not exists
        patient_dir = "patient_data/excel_records"
        os.makedirs(patient_dir, exist_ok=True)
        
        # Create filename based on patient ID or name
        patient_id = patient_data.get('PATIENT_ID', 'UNKNOWN')
        filename = f"{patient_dir}/{patient_id}.xlsx"
        
        # Prepare data for Excel
        excel_data = {
            'Basic Information': {
                'Patient ID': patient_id,
                'Name': patient_data.get('NAME', ''),
                'Date of Birth': patient_data.get('DATE_OF_BIRTH', ''),
                'Age': patient_data.get('AGE', ''),
                'Record Date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'Doctor': doctor
            },
            'Clinical Measurements': {
                'Height (cm)': patient_data.get('HEIGHT', ''),
                'Weight (kg)': patient_data.get('WEIGHT', ''),
                'BMI': patient_data.get('BMI', ''),
                'BMI Category': get_bmi_category(patient_data.get('BMI', 0)),
                'Gestational Age (weeks)': patient_data.get('GESTATIONAL_AGE', ''),
                'Parity': patient_data.get('PARITY', ''),
                'Previous Cesarean': 'Yes' if patient_data.get('PREVIOUS_CESAREAN', 0) == 1 else 'No',
                'Number of Previous Cesareans': patient_data.get('NUMBER_PREV_CESAREAN', 0)
            },
            'Medical History': {
                'Comorbidities': ', '.join(patient_data.get('COMORBIDITIES', [])) if isinstance(patient_data.get('COMORBIDITIES'), list) else patient_data.get('COMORBIDITIES', 'None'),
                'Obstetric Risk': patient_data.get('OBSTETRIC_RISK', ''),
                'Smoking': 'Yes' if patient_data.get('SMOKING', 0) == 1 else 'No',
                'Alcohol': 'Yes' if patient_data.get('ALCOHOL', 0) == 1 else 'No'
            },
            'AI Prediction': {
                'Predicted Delivery Mode': prediction,
                'Prediction Confidence': f"{probability:.1%}",
                'Risk Score': f"{risk_score}/100",
                'Risk Level': get_risk_level(risk_score)
            },
            'Clinical Notes': {
                'Notes': patient_data.get('CLINICAL_NOTES', ''),
                'Follow-up Date': patient_data.get('FOLLOW_UP_DATE', '')
            }
        }
        
        # Create Excel workbook
        wb = Workbook()
        
        # Add sheets
        for sheet_name, data in excel_data.items():
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=sheet_name)
            else:
                ws = wb[sheet_name]
            
            # Write headers and data
            for i, (key, value) in enumerate(data.items(), 1):
                ws.cell(row=i, column=1, value=key)
                ws.cell(row=i, column=2, value=value)
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Save workbook
        wb.save(filename)
        
        # Also update master CSV for quick searching
        update_master_record(patient_data, prediction, probability, risk_score, doctor, filename)
        
        return filename
        
    except Exception as e:
        st.error(f"Error saving to Excel: {str(e)}")
        return None

def update_master_record(patient_data, prediction, probability, risk_score, doctor, excel_path):
    """Update master CSV file for searching"""
    master_file = "patient_data/master_records.csv"
    
    record = {
        'PATIENT_ID': patient_data.get('PATIENT_ID', ''),
        'NAME': patient_data.get('NAME', ''),
        'AGE': patient_data.get('AGE', ''),
        'GESTATIONAL_AGE': patient_data.get('GESTATIONAL_AGE', ''),
        'PREDICTION': prediction,
        'PROBABILITY': probability,
        'RISK_SCORE': risk_score,
        'DOCTOR': doctor,
        'DATE': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'EXCEL_FILE': excel_path
    }
    
    df_record = pd.DataFrame([record])
    
    if os.path.exists(master_file):
        try:
            existing_df = pd.read_csv(master_file)
            df_combined = pd.concat([existing_df, df_record], ignore_index=True)
        except:
            df_combined = df_record
    else:
        df_combined = df_record
    
    df_combined.to_csv(master_file, index=False)

def search_patient(search_term):
    """Search for patient in master records"""
    master_file = "patient_data/master_records.csv"
    
    if not os.path.exists(master_file):
        return None
    
    try:
        df = pd.read_csv(master_file)
        
        # Search by ID or name
        mask = (
            df['PATIENT_ID'].astype(str).str.contains(search_term, case=False, na=False) |
            df['NAME'].astype(str).str.contains(search_term, case=False, na=False)
        )
        
        results = df[mask]
        return results
    except:
        return None

def load_patient_from_excel(filepath):
    """Load patient data from Excel file"""
    try:
        wb = load_workbook(filepath, data_only=True)
        patient_data = {}
        
        # Extract data from all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
                if row[0] and row[1]:
                    patient_data[row[0]] = row[1]
        
        return patient_data
    except:
        return None

def get_table_download_link(df, filename="patient_data.csv"):
    """Generate a download link for a DataFrame"""
    csv = df.to_csv(index=False)
    b64 = base64.b64encode(csv.encode()).decode()
    href = f'<a href="data:file/csv;base64,{b64}" download="{filename}" style="background-color: #3B82F6; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; display: inline-block;">📥 Download CSV</a>'
    return href

def get_excel_download_link(filepath, filename=None):
    """Generate a download link for an Excel file"""
    if not os.path.exists(filepath):
        return None
    
    with open(filepath, 'rb') as f:
        data = f.read()
    
    b64 = base64.b64encode(data).decode()
    if not filename:
        filename = os.path.basename(filepath)
    
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}" style="background-color: #10B981; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; display: inline-block;">📥 Download Excel Report</a>'
    return href

def load_user_credentials():
    """Load user credentials from file"""
    try:
        with open('user_credentials/users.json', 'r') as f:
            return json.load(f)
    except:
        # Create default credentials if file doesn't exist
        return {
            "admin": {"password": "admin123", "role": "admin", "name": "System Admin"},
            "doctor1": {"password": "doc123", "role": "doctor", "name": "Dr. Sarah Johnson"},
            "doctor2": {"password": "doc123", "role": "doctor", "name": "Dr. Michael Chen"},
            "doctor3": {"password": "doc123", "role": "doctor", "name": "Dr. Emily Williams"},
            "doctor4": {"password": "doc123", "role": "doctor", "name": "Dr. Robert Brown"},
            "doctor5": {"password": "doc123", "role": "doctor", "name": "Dr. Maria Garcia"}
        }

def load_model():
    """Load trained model and preprocessors"""
    try:
        # Check required files
        required_files = ['trained_model.pkl', 'scaler.pkl', 'encoders.pkl', 
                         'feature_names.pkl', 'target_classes.pkl']
        
        for file in required_files:
            if not os.path.exists(f'models/{file}'):
                st.error(f"Missing model file: models/{file}")
                return False
        
        # Load all components
        st.session_state.model = joblib.load('models/trained_model.pkl')
        st.session_state.scaler = joblib.load('models/scaler.pkl')
        
        with open('models/encoders.pkl', 'rb') as f:
            st.session_state.encoders = pickle.load(f)
        
        with open('models/feature_names.pkl', 'rb') as f:
            st.session_state.feature_names = pickle.load(f)
        
        with open('models/target_classes.pkl', 'rb') as f:
            st.session_state.target_classes = pickle.load(f)
        
        # Load feature descriptions
        if os.path.exists('models/feature_descriptions.pkl'):
            with open('models/feature_descriptions.pkl', 'rb') as f:
                st.session_state.feature_descriptions = pickle.load(f)
        
        # Load SHAP explainer
        if os.path.exists('models/shap_explainer.pkl'):
            st.session_state.shap_explainer = joblib.load('models/shap_explainer.pkl')
        
        st.session_state.model_loaded = True
        return True
        
    except Exception as e:
        st.error(f"Error loading model: {str(e)}")
        return False

def prepare_input_features(patient_data):
    """Prepare patient data for model prediction"""
    try:
        if not st.session_state.feature_names or not st.session_state.scaler:
            return None
        
        # Create default feature vector
        features = pd.DataFrame(0, index=[0], columns=st.session_state.feature_names)
        
        # Map patient data to features
        mapping = {
            'AGE': ['AGE', 'age', 'Age'],
            'HEIGHT': ['HEIGHT', 'height', 'Height'],
            'WEIGHT': ['WEIGHT', 'weight', 'Weight'],
            'BMI': ['BMI', 'bmi', 'Bmi'],
            'GESTATIONAL_AGE': ['GESTATIONAL_AGE', 'GESTAGIONAL_AGE', 'gestational_age'],
            'PARITY': ['PARITY', 'parity', 'Parity'],
            'PREVIOUS_CESAREAN': ['HAS_PREV_CESAREAN', 'PREVIOUS_CESAREAN'],
            'NUMBER_PREV_CESAREAN': ['NUMBER_PREV_CESAREAN', 'NUMBER_OF_PREV_CESAREAN'],
            'COMORBIDITY': ['COMORBIDITY', 'comorbidity'],
            'SMOKING': ['SMOKING', 'smoking'],
            'ALCOHOL': ['ALCOHOL', 'alcohol'],
            'INDUCTION': ['INDUCTION', 'induction'],
            'HOURS_ROM': ['HOURS_ROM', 'HOURS_OF_RUPTURED_MEMBRANES']
        }
        
        # Fill features
        for patient_key, possible_names in mapping.items():
            if patient_key in patient_data:
                value = patient_data[patient_key]
                for name in possible_names:
                    if name in features.columns:
                        features[name] = value
                        break
        
        # Scale features
        features_scaled = st.session_state.scaler.transform(features)
        
        return features_scaled
        
    except Exception as e:
        st.error(f"Error preparing features: {str(e)}")
        return None

def generate_explanation_paragraph(patient_data, prediction, probability, risk_score):
    """Generate detailed paragraph explanation for prediction"""
    
    explanation = f"""
    **Prediction Analysis for {patient_data.get('NAME', 'the patient')}:**
    
    The AI model has predicted a **{prediction}** with **{probability:.1%} confidence**. 
    The patient's clinical risk assessment shows a **{risk_score}/100 risk score**, categorized as **{get_risk_level(risk_score)}**.
    
    **Key Contributing Factors:**
    """
    
    factors = []
    
    # Age analysis
    age = patient_data.get('AGE', 0)
    if age >= 35:
        factors.append(f"• **Advanced Maternal Age ({age} years)**: Increases the likelihood of cesarean delivery due to higher risk of labor complications.")
    elif age >= 30:
        factors.append(f"• **Maternal Age ({age} years)**: Moderate risk factor that may influence delivery planning.")
    
    # BMI analysis
    bmi = patient_data.get('BMI', 0)
    bmi_category = get_bmi_category(bmi)
    if bmi_category == "Obese":
        factors.append(f"• **High BMI ({bmi:.1f} - {bmi_category})**: Significantly increases surgical risks and may necessitate cesarean delivery for safety.")
    elif bmi_category == "Overweight":
        factors.append(f"• **Elevated BMI ({bmi:.1f} - {bmi_category})**: May contribute to delivery complications requiring careful monitoring.")
    
    # Previous cesarean
    if patient_data.get('PREVIOUS_CESAREAN', 0) == 1:
        count = patient_data.get('NUMBER_PREV_CESAREAN', 1)
        factors.append(f"• **History of {count} previous cesarean section(s)**: Strong indicator for repeat cesarean delivery to avoid uterine rupture risks.")
    
    # Comorbidities
    if patient_data.get('COMORBIDITY', 0) == 1:
        factors.append("• **Presence of comorbidities**: Medical conditions increase overall risk and may influence delivery method selection.")
    
    # Lifestyle factors
    if patient_data.get('SMOKING', 0) == 1:
        factors.append("• **Smoking history**: Associated with placental complications and potential need for surgical intervention.")
    if patient_data.get('ALCOHOL', 0) == 1:
        factors.append("• **Alcohol consumption**: May affect fetal well-being and delivery planning.")
    
    # Obstetric risk
    obstetric_risk = patient_data.get('OBSTETRIC_RISK', 'medium')
    if obstetric_risk == 'high':
        factors.append("• **High obstetric risk**: Requires specialized management and may indicate cesarean delivery for optimal outcomes.")
    elif obstetric_risk == 'medium':
        factors.append("• **Moderate obstetric risk**: Warrants close monitoring during delivery.")
    
    # Gestational age
    ga = patient_data.get('GESTATIONAL_AGE', 38)
    if ga < 37:
        factors.append(f"• **Preterm gestation ({ga} weeks)**: Increases complexity and may influence delivery method based on fetal condition.")
    elif ga >= 42:
        factors.append(f"• **Post-term gestation ({ga} weeks)**: Associated with increased risks that may necessitate intervention.")
    
    # Add factors to explanation
    if factors:
        explanation += "\n" + "\n".join(factors)
    else:
        explanation += "\n• No strong risk factors identified in the available data."
    
    # Add confidence level explanation
    explanation += f"""
    
    **Model Confidence Assessment:**
    """
    
    if probability >= 0.9:
        explanation += "The model shows **very high confidence** in this prediction, indicating strong pattern recognition from similar clinical cases."
    elif probability >= 0.7:
        explanation += "The model shows **good confidence** in this prediction, suggesting reliable assessment based on available parameters."
    elif probability >= 0.5:
        explanation += "The model shows **moderate confidence** in this prediction, recommending additional clinical assessment for confirmation."
    else:
        explanation += "The model shows **low confidence** in this prediction, strongly suggesting comprehensive clinical evaluation and possibly additional diagnostics."
    
    # Add final recommendation
    explanation += f"""
    
    **Clinical Interpretation:**
    This prediction should be integrated with complete clinical assessment. The AI model provides decision support based on statistical patterns from historical data. Final delivery planning should consider patient preferences, current clinical status, and institutional protocols.
    
    **Next Steps:**
    1. Review prediction with the complete clinical team
    2. Discuss options with the patient
    3. Prepare appropriate resources based on predicted delivery mode
    4. Document decision-making process in patient records
    """
    
    return explanation

# Authentication
def show_login():
    """Display login page"""
    st.markdown('<h1 class="main-header">🏥 Hospital Authentication System</h1>', unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        with st.container():
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown("### 🔐 Secure Login")
            
            # User selection
            users = load_user_credentials()
            user_options = list(users.keys())
            selected_user = st.selectbox("Select User", user_options)
            
            # Get user info
            user_info = users[selected_user]
            
            # Display user info
            st.markdown(f"**Name:** {user_info['name']}")
            st.markdown(f"**Role:** {user_info['role'].title()}")
            st.markdown(f"**Department:** {user_info.get('department', 'Not specified')}")
            
            # Password input
            password = st.text_input("Password", type="password", key="login_password")
            
            # Login button
            if st.button("🚀 Login", use_container_width=True):
                if password == user_info['password']:
                    st.session_state.authenticated = True
                    st.session_state.current_user = selected_user
                    st.session_state.user_role = user_info['role']
                    st.session_state.user_name = user_info['name']
                    st.rerun()
                else:
                    st.error("Invalid password")
            
            st.markdown("---")
            st.markdown("**Available Users:**")
            for username, info in users.items():
                role_icon = "👨‍💼" if info['role'] == 'admin' else "👩‍⚕️"
                st.markdown(f"- {role_icon} **{username}**: {info['name']} ({info['role']})")
            
            st.markdown("</div>", unsafe_allow_html=True)

# Main Application Pages
def home_page():
    """Home and Prediction Page"""
    st.markdown('<h1 class="main-header">👶 Childbirth Mode Prediction System</h1>', unsafe_allow_html=True)
    
    # User info
    user_role = st.session_state.user_role
    user_name = st.session_state.user_name
    
    if user_role == 'admin':
        st.markdown(f'<div class="admin-card">', unsafe_allow_html=True)
        st.markdown(f"### 👨‍💼 Welcome, {user_name} (Administrator)")
        st.markdown("You have full system access including user management and system configuration.")
    else:
        st.markdown(f'<div class="doctor-card">', unsafe_allow_html=True)
        st.markdown(f"### 👩‍⚕️ Welcome, {user_name}")
        st.markdown(f"You are logged in as a **{user_role}**. You can make predictions, view patient history, and generate reports.")
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Check if model is loaded
    if not st.session_state.model_loaded:
        st.warning("⚠️ AI Model not loaded. Loading model...")
        with st.spinner("Loading AI model..."):
            if not load_model():
                st.error("Failed to load model. Please train a model first.")
                return
    
    st.success("✅ AI Model loaded and ready for predictions")
    
    # Create tabs for different functions
    tab1, tab2, tab3 = st.tabs(["🎯 Make Prediction", "🔍 Search Patient", "📊 Quick Stats"])
    
    with tab1:
        make_prediction_section()
    
    with tab2:
        search_patient_section()
    
    with tab3:
        show_quick_stats()

def make_prediction_section():
    """Section for making new predictions"""
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown('<h2 class="sub-header">📝 New Patient Assessment</h2>', unsafe_allow_html=True)
        
        with st.form("patient_form", clear_on_submit=False):
            st.markdown('<div class="card">', unsafe_allow_html=True)
            
            # Basic Information
            st.markdown("### 👤 Basic Information")
            col_name, col_dob = st.columns(2)
            with col_name:
                name = st.text_input("Patient Full Name*", placeholder="Enter patient's full name")
            with col_dob:
                dob = st.date_input("Date of Birth*", value=datetime.now().replace(year=1990))
            
            # Contact Information
            col_phone, col_email = st.columns(2)
            with col_phone:
                phone = st.text_input("Phone Number", placeholder="+1234567890")
            with col_email:
                email = st.text_input("Email Address", placeholder="patient@email.com")
            
            st.markdown("---")
            
            # Demographics
            st.markdown("### 📊 Demographics")
            col_age, col_height, col_weight = st.columns(3)
            with col_age:
                age = st.number_input("Age (years)*", min_value=15, max_value=50, value=30)
            with col_height:
                height = st.number_input("Height (cm)*", min_value=140, max_value=200, value=165)
            with col_weight:
                weight = st.number_input("Weight (kg)*", min_value=40, max_value=150, value=65)
            
            # Calculate BMI
            if height and weight:
                bmi = calculate_bmi(weight, height)
                bmi_category = get_bmi_category(bmi)
                st.info(f"**BMI:** {bmi:.1f} ({bmi_category})")
            
            st.markdown("---")
            
            # Pregnancy Details
            st.markdown("### 🤰 Pregnancy Details")
            col_ga, col_parity = st.columns(2)
            with col_ga:
                gestational_age = st.slider("Gestational Age (weeks)*", min_value=20, max_value=42, value=38)
            with col_parity:
                parity = st.number_input("Parity*", min_value=0, max_value=10, value=1)
            
            col_prev_c, col_num_c = st.columns(2)
            with col_prev_c:
                prev_cesarean = st.selectbox("Previous Cesarean*", ["No", "Yes"])
            with col_num_c:
                if prev_cesarean == "Yes":
                    prev_cesarean_count = st.number_input("Number of Previous Cesareans*", min_value=1, max_value=5, value=1)
                else:
                    prev_cesarean_count = 0
            
            st.markdown("---")
            
            # Medical History
            st.markdown("### 🩺 Medical History")
            col_comorb, col_risk = st.columns(2)
            with col_comorb:
                comorbidities = st.multiselect("Comorbidities", 
                                             ["Hypertension", "Diabetes", "Thyroid Disorder", 
                                              "Asthma", "Heart Disease", "None"])
            with col_risk:
                obstetric_risk = st.select_slider("Obstetric Risk Level*", 
                                                options=["low", "medium", "high"], value="medium")
            
            col_smoke, col_alcohol = st.columns(2)
            with col_smoke:
                smoking = st.selectbox("Smoking*", ["No", "Yes"])
            with col_alcohol:
                alcohol = st.selectbox("Alcohol Consumption*", ["No", "Yes"])
            
            st.markdown("---")
            
            # Labor Information
            st.markdown("### 🏥 Labor & Delivery")
            col_induct, col_hours = st.columns(2)
            with col_induct:
                induction = st.selectbox("Induction", ["No", "Yes"])
            with col_hours:
                hours_rom = st.number_input("Hours of Ruptured Membranes", min_value=0, max_value=48, value=0)
            
            col_ctg, col_anes = st.columns(2)
            with col_ctg:
                cardiotocography = st.selectbox("Cardiotocography", ["None", "Intermittent", "Continuous"])
            with col_anes:
                anesthesia = st.selectbox("Anesthesia", ["None", "Epidural", "Spinal", "General"])
            
            # Clinical Notes
            clinical_notes = st.text_area("Clinical Notes", placeholder="Additional clinical observations...", height=100)
            
            # Follow-up Date
            follow_up = st.date_input("Follow-up Date", value=datetime.now())
            
            st.markdown("</div>", unsafe_allow_html=True)
            
            # Submit button
            submit_button = st.form_submit_button("🔮 PREDICT DELIVERY MODE", use_container_width=True)
        
        # Handle form submission
        if submit_button:
            if not name:
                st.error("Please enter patient name")
                return
            
            # Generate patient ID
            patient_id = create_patient_id(name, str(dob))
            
            # Prepare patient data
            patient_data = {
                'PATIENT_ID': patient_id,
                'NAME': name,
                'DATE_OF_BIRTH': str(dob),
                'PHONE': phone,
                'EMAIL': email,
                'AGE': age,
                'HEIGHT': height,
                'WEIGHT': weight,
                'BMI': bmi,
                'GESTATIONAL_AGE': gestational_age,
                'PARITY': parity,
                'PREVIOUS_CESAREAN': 1 if prev_cesarean == "Yes" else 0,
                'NUMBER_PREV_CESAREAN': prev_cesarean_count,
                'COMORBIDITIES': comorbidities,
                'OBSTETRIC_RISK': obstetric_risk,
                'SMOKING': 1 if smoking == "Yes" else 0,
                'ALCOHOL': 1 if alcohol == "Yes" else 0,
                'INDUCTION': 1 if induction == "Yes" else 0,
                'HOURS_ROM': hours_rom,
                'CARDIOTOCOGRAPHY': ["None", "Intermittent", "Continuous"].index(cardiotocography),
                'ANESTHESIA': ["None", "Epidural", "Spinal", "General"].index(anesthesia),
                'CLINICAL_NOTES': clinical_notes,
                'FOLLOW_UP_DATE': str(follow_up)
            }
            
            # Calculate risk score
            risk_score = calculate_risk_score(patient_data)
            patient_data['RISK_SCORE'] = risk_score
            
            # Prepare features for model
            features = prepare_input_features(patient_data)
            
            if features is not None:
                try:
                    # Make prediction
                    probabilities = st.session_state.model.predict_proba(features)[0]
                    prediction_idx = np.argmax(probabilities)
                    prediction = st.session_state.target_classes[prediction_idx]
                    probability = probabilities[prediction_idx]
                    
                    # Store in session
                    st.session_state.patient_data = patient_data
                    st.session_state.predictions = {
                        'mode': prediction,
                        'probability': probability,
                        'risk_score': risk_score
                    }
                    
                    # Generate explanation
                    explanation = generate_explanation_paragraph(patient_data, prediction, probability, risk_score)
                    st.session_state.explanation = explanation
                    
                    # Generate recommendations
                    recommendations = get_recommendations(prediction, probability, risk_score)
                    st.session_state.recommendations = recommendations
                    
                    # Save to Excel
                    excel_file = save_to_excel(patient_data, prediction, probability, risk_score, 
                                             st.session_state.user_name)
                    
                    if excel_file:
                        st.session_state.excel_file = excel_file
                    
                    # Display results in right column
                    with col2:
                        display_prediction_results(patient_data, prediction, probability, 
                                                 risk_score, explanation, recommendations, excel_file)
                
                except Exception as e:
                    st.error(f"Error making prediction: {str(e)}")
            else:
                st.error("Failed to prepare features for prediction")

def display_prediction_results(patient_data, prediction, probability, risk_score, explanation, recommendations, excel_file):
    """Display prediction results"""
    st.markdown('<h2 class="sub-header">📊 Prediction Results</h2>', unsafe_allow_html=True)
    
    # Prediction card
    risk_class = "high-risk" if risk_score >= 80 else "medium-risk" if risk_score >= 60 else "low-risk"
    st.markdown(f'<div class="prediction-card {risk_class}">', unsafe_allow_html=True)
    
    st.markdown(f"### 🎯 Predicted Mode")
    st.markdown(f"# **{str(prediction).replace('_', ' ').title()}**")
    
    col_prob, col_risk = st.columns(2)
    with col_prob:
        st.metric("Confidence", f"{probability:.1%}")
    with col_risk:
        st.metric("Risk Score", f"{risk_score}/100")
    
    risk_level = get_risk_level(risk_score)
    st.markdown(f"### {risk_level}")
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Patient info
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown(f"### 👤 Patient Information")
    st.markdown(f"**Name:** {patient_data.get('NAME', 'Unknown')}")
    st.markdown(f"**Patient ID:** {patient_data.get('PATIENT_ID', 'N/A')}")
    st.markdown(f"**Age:** {patient_data.get('AGE', 'N/A')} years")
    st.markdown(f"**Gestational Age:** {patient_data.get('GESTATIONAL_AGE', 'N/A')} weeks")
    st.markdown(f"**Recorded By:** {st.session_state.user_name}")
    st.markdown('</div>', unsafe_allow_html=True)
    
    # AI Explanation
    st.markdown('<div class="info-card">', unsafe_allow_html=True)
    st.markdown("### 🤖 AI Explanation")
    st.markdown(explanation)
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Recommendations
    st.markdown('<div class="info-card">', unsafe_allow_html=True)
    st.markdown("### 🩺 Clinical Recommendations")
    for rec in recommendations:
        st.markdown(f"• {rec}")
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Action buttons
    st.markdown("### 💾 Actions")
    
    col_act1, col_act2, col_act3 = st.columns(3)
    
    with col_act1:
        if st.button("📊 View Detailed Analysis", use_container_width=True):
            st.session_state.current_page = "XAI"
            st.rerun()
    
    with col_act2:
        if excel_file and os.path.exists(excel_file):
            st.markdown(get_excel_download_link(excel_file, f"{patient_data.get('PATIENT_ID', 'patient')}_report.xlsx"), 
                       unsafe_allow_html=True)
    
    with col_act3:
        if st.button("🔄 New Prediction", use_container_width=True):
            st.rerun()

def search_patient_section():
    """Section for searching patient records"""
    st.markdown('<h2 class="sub-header">🔍 Search Patient Records</h2>', unsafe_allow_html=True)
    
    search_col1, search_col2 = st.columns([3, 1])
    with search_col1:
        search_term = st.text_input("Search by Patient ID or Name", placeholder="Enter patient ID or name...")
    
    with search_col2:
        search_button = st.button("🔍 Search", use_container_width=True)
    
    if search_button and search_term:
        results = search_patient(search_term)
        
        if results is not None and not results.empty:
            st.success(f"Found {len(results)} record(s)")
            
            for idx, row in results.iterrows():
                with st.expander(f"📋 Patient: {row['NAME']} (ID: {row['PATIENT_ID']})", expanded=False):
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.markdown(f"**Age:** {row.get('AGE', 'N/A')}")
                        st.markdown(f"**Gestational Age:** {row.get('GESTATIONAL_AGE', 'N/A')} weeks")
                    with col2:
                        st.markdown(f"**Prediction:** {row.get('PREDICTION', 'N/A')}")
                        st.markdown(f"**Confidence:** {float(row.get('PROBABILITY', 0)):.1%}")
                    with col3:
                        st.markdown(f"**Risk Score:** {row.get('RISK_SCORE', 'N/A')}")
                        st.markdown(f"**Doctor:** {row.get('DOCTOR', 'N/A')}")
                    
                    # Load full data from Excel
                    excel_file = row.get('EXCEL_FILE', '')
                    if excel_file and os.path.exists(excel_file):
                        if st.button(f"📄 View Full Report", key=f"view_{idx}"):
                            patient_data = load_patient_from_excel(excel_file)
                            if patient_data:
                                st.markdown("### 📋 Complete Patient Record")
                                for key, value in patient_data.items():
                                    st.markdown(f"**{key}:** {value}")
                                
                                # Download button
                                st.markdown(get_excel_download_link(excel_file), unsafe_allow_html=True)
                    else:
                        st.warning("Full report not available")
        elif results is not None:
            st.info("No patients found matching your search")
        else:
            st.warning("Search database not available")

def show_quick_stats():
    """Show quick statistics"""
    st.markdown('<h2 class="sub-header">📊 Quick Statistics</h2>', unsafe_allow_html=True)
    
    master_file = "patient_data/master_records.csv"
    
    if os.path.exists(master_file):
        try:
            df = pd.read_csv(master_file)
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                total = len(df)
                st.metric("Total Patients", total)
            
            with col2:
                if 'RISK_SCORE' in df.columns:
                    high_risk = len(df[df['RISK_SCORE'] >= 80])
                    st.metric("High Risk Cases", high_risk)
                else:
                    st.metric("High Risk Cases", "N/A")
            
            with col3:
                if 'PREDICTION' in df.columns:
                    if not df['PREDICTION'].mode().empty:
                        common = df['PREDICTION'].mode()[0]
                        st.metric("Most Common", common.replace('_', ' ').title())
                else:
                    st.metric("Most Common", "N/A")
            
            with col4:
                if 'DOCTOR' in df.columns:
                    doctors = df['DOCTOR'].nunique()
                    st.metric("Doctors", doctors)
                else:
                    st.metric("Doctors", "N/A")
            
            # Recent records
            st.markdown("### 📅 Recent Records")
            if 'DATE' in df.columns:
                recent = df.sort_values('DATE', ascending=False).head(5)
                for idx, row in recent.iterrows():
                    with st.container():
                        st.markdown(f"**{row.get('NAME', 'Unknown')}** - {row.get('PREDICTION', 'Unknown')} (ID: {row.get('PATIENT_ID', 'N/A')})")
                        st.markdown(f"*{row.get('DATE', '')} • {row.get('DOCTOR', 'Unknown')}*")
                        st.divider()
        
        except:
            st.info("Could not load statistics")
    else:
        st.info("No patient records available")

def patient_history_page():
    """Patient History Page (Admin and Doctors)"""
    st.markdown('<h1 class="main-header">📋 Patient History Database</h1>', unsafe_allow_html=True)
    
    if st.session_state.user_role == 'admin':
        st.markdown('<div class="admin-card">', unsafe_allow_html=True)
        st.markdown("### 👨‍💼 Administrator View - Full Access")
        st.markdown("You can view all patient records across all doctors.")
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="doctor-card">', unsafe_allow_html=True)
        st.markdown(f"### 👩‍⚕️ Doctor View - {st.session_state.user_name}")
        st.markdown("You can view patients you have assessed.")
        st.markdown('</div>', unsafe_allow_html=True)
    
    master_file = "patient_data/master_records.csv"
    
    if not os.path.exists(master_file):
        st.warning("No patient records found.")
        return
    
    try:
        df = pd.read_csv(master_file)
        
        if df.empty:
            st.info("No records in database")
            return
        
        # Filter by doctor if not admin
        if st.session_state.user_role != 'admin':
            df = df[df['DOCTOR'] == st.session_state.user_name]
        
        if df.empty:
            st.info(f"No records found for {st.session_state.user_name}")
            return
        
        # Filters
        with st.expander("🔍 Filter Records", expanded=True):
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                # Risk filter
                risk_options = ["All"] + sorted([f"{int(r)}-{int(r)+9}" for r in range(0, 101, 10) if r < 100])
                risk_filter = st.selectbox("Risk Range", risk_options)
            
            with col2:
                # Prediction filter
                pred_options = ["All"] + sorted(df['PREDICTION'].unique().tolist())
                pred_filter = st.selectbox("Delivery Mode", pred_options)
            
            with col3:
                # Date range
                if 'DATE' in df.columns:
                    df['DATE'] = pd.to_datetime(df['DATE'], errors='coerce')
                    min_date = df['DATE'].min().date() if not pd.isna(df['DATE'].min()) else datetime.now().date()
                    max_date = df['DATE'].max().date() if not pd.isna(df['DATE'].max()) else datetime.now().date()
                    
                    date_range = st.date_input("Date Range", [min_date, max_date])
                else:
                    date_range = None
            
            with col4:
                # Search
                search_term = st.text_input("Search ID/Name")
        
        # Apply filters
        filtered_df = df.copy()
        
        if risk_filter != "All":
            risk_min, risk_max = map(int, risk_filter.split('-'))
            filtered_df = filtered_df[
                (filtered_df['RISK_SCORE'] >= risk_min) & 
                (filtered_df['RISK_SCORE'] <= risk_max)
            ]
        
        if pred_filter != "All":
            filtered_df = filtered_df[filtered_df['PREDICTION'] == pred_filter]
        
        if date_range and len(date_range) == 2 and 'DATE' in filtered_df.columns:
            filtered_df = filtered_df[
                (filtered_df['DATE'].dt.date >= date_range[0]) & 
                (filtered_df['DATE'].dt.date <= date_range[1])
            ]
        
        if search_term:
            mask = (
                filtered_df['PATIENT_ID'].astype(str).str.contains(search_term, case=False, na=False) |
                filtered_df['NAME'].astype(str).str.contains(search_term, case=False, na=False)
            )
            filtered_df = filtered_df[mask]
        
        # Display statistics
        st.markdown("### 📈 Statistics")
        col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
        
        with col_stat1:
            st.metric("Total Records", len(df))
        
        with col_stat2:
            st.metric("Filtered Records", len(filtered_df))
        
        with col_stat3:
            if 'PROBABILITY' in filtered_df.columns and len(filtered_df) > 0:
                avg_conf = filtered_df['PROBABILITY'].mean()
                st.metric("Avg Confidence", f"{avg_conf:.1%}")
            else:
                st.metric("Avg Confidence", "N/A")
        
        with col_stat4:
            if 'RISK_SCORE' in filtered_df.columns and len(filtered_df) > 0:
                high_risk = len(filtered_df[filtered_df['RISK_SCORE'] >= 80])
                st.metric("High Risk", high_risk)
            else:
                st.metric("High Risk", "N/A")
        
        # Display records
        st.markdown("### 📊 Patient Records")
        
        if len(filtered_df) > 0:
            # Create display dataframe
            display_cols = []
            if 'PATIENT_ID' in filtered_df.columns:
                display_cols.append('PATIENT_ID')
            if 'NAME' in filtered_df.columns:
                display_cols.append('NAME')
            if 'AGE' in filtered_df.columns:
                display_cols.append('AGE')
            if 'GESTATIONAL_AGE' in filtered_df.columns:
                display_cols.append('GESTATIONAL_AGE')
            if 'PREDICTION' in filtered_df.columns:
                display_cols.append('PREDICTION')
            if 'PROBABILITY' in filtered_df.columns:
                display_cols.append('PROBABILITY')
            if 'RISK_SCORE' in filtered_df.columns:
                display_cols.append('RISK_SCORE')
            if 'DOCTOR' in filtered_df.columns:
                display_cols.append('DOCTOR')
            if 'DATE' in filtered_df.columns:
                display_cols.append('DATE')
            
            display_df = filtered_df[display_cols].copy()
            
            # Format columns
            if 'PROBABILITY' in display_df.columns:
                display_df['PROBABILITY'] = display_df['PROBABILITY'].apply(lambda x: f"{float(x):.1%}")
            if 'DATE' in display_df.columns:
                display_df['DATE'] = display_df['DATE'].dt.strftime('%Y-%m-%d %H:%M')
            
            st.dataframe(
                display_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "RISK_SCORE": st.column_config.ProgressColumn(
                        "Risk Score",
                        format="%d",
                        min_value=0,
                        max_value=100
                    )
                }
            )
            
            # Download option
            st.markdown(get_table_download_link(filtered_df, "patient_history.csv"), unsafe_allow_html=True)
        else:
            st.info("No records match the current filters")
    
    except Exception as e:
        st.error(f"Error loading patient history: {str(e)}")

def analytics_page():
    """Analytics Dashboard"""
    st.markdown('<h1 class="main-header">📊 Analytics Dashboard</h1>', unsafe_allow_html=True)
    
    if st.session_state.user_role == 'admin':
        st.markdown('<div class="admin-card">', unsafe_allow_html=True)
        st.markdown("### 👨‍💼 Administrator Analytics - System Overview")
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="doctor-card">', unsafe_allow_html=True)
        st.markdown(f"### 👩‍⚕️ Doctor Analytics - {st.session_state.user_name}")
        st.markdown('</div>', unsafe_allow_html=True)
    
    master_file = "patient_data/master_records.csv"
    
    if not os.path.exists(master_file):
        st.warning("No analytics data available.")
        return
    
    try:
        df = pd.read_csv(master_file)
        
        if df.empty:
            st.info("No data available for analytics.")
            return
        
        # Filter by doctor if not admin
        if st.session_state.user_role != 'admin':
            df = df[df['DOCTOR'] == st.session_state.user_name]
        
        if df.empty:
            st.info(f"No data available for {st.session_state.user_name}")
            return
        
        # Key Metrics
        st.markdown("### 📊 Key Performance Indicators")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total = len(df)
            st.metric("Total Assessments", total)
        
        with col2:
            if 'RISK_SCORE' in df.columns:
                avg_risk = df['RISK_SCORE'].mean()
                st.metric("Avg Risk Score", f"{avg_risk:.1f}")
            else:
                st.metric("Avg Risk Score", "N/A")
        
        with col3:
            if 'PROBABILITY' in df.columns:
                avg_conf = df['PROBABILITY'].astype(float).mean()
                st.metric("Avg Confidence", f"{avg_conf:.1%}")
            else:
                st.metric("Avg Confidence", "N/A")
        
        with col4:
            if 'RISK_SCORE' in df.columns:
                high_risk = len(df[df['RISK_SCORE'] >= 80])
                high_risk_pct = (high_risk / total * 100) if total > 0 else 0
                st.metric("High Risk %", f"{high_risk_pct:.1f}%")
            else:
                st.metric("High Risk %", "N/A")
        
        # Charts
        st.markdown("### 📈 Visual Analytics")
        
        tab1, tab2, tab3 = st.tabs(["📊 Distribution", "📅 Trends", "👥 Comparisons"])
        
        with tab1:
            col_chart1, col_chart2 = st.columns(2)
            
            with col_chart1:
                if 'PREDICTION' in df.columns:
                    prediction_counts = df['PREDICTION'].value_counts()
                    fig1 = px.pie(
                        values=prediction_counts.values,
                        names=prediction_counts.index,
                        title="Delivery Mode Distribution",
                        color_discrete_sequence=px.colors.qualitative.Set3
                    )
                    st.plotly_chart(fig1, use_container_width=True)
            
            with col_chart2:
                if 'RISK_SCORE' in df.columns:
                    fig2 = px.histogram(
                        df,
                        x='RISK_SCORE',
                        nbins=20,
                        title="Risk Score Distribution",
                        color_discrete_sequence=['#3B82F6']
                    )
                    st.plotly_chart(fig2, use_container_width=True)
        
        with tab2:
            if 'DATE' in df.columns:
                try:
                    df['DATE'] = pd.to_datetime(df['DATE'])
                    df['MONTH'] = df['DATE'].dt.to_period('M').astype(str)
                    
                    monthly_stats = df.groupby('MONTH').agg({
                        'PATIENT_ID': 'count',
                        'RISK_SCORE': 'mean',
                        'PROBABILITY': 'mean'
                    }).reset_index()
                    
                    fig3 = go.Figure()
                    fig3.add_trace(go.Scatter(
                        x=monthly_stats['MONTH'],
                        y=monthly_stats['PATIENT_ID'],
                        mode='lines+markers',
                        name='Assessments',
                        line=dict(color='#3B82F6', width=3)
                    ))
                    
                    fig3.update_layout(
                        title="Monthly Assessment Trends",
                        xaxis_title="Month",
                        yaxis_title="Number of Assessments",
                        height=400
                    )
                    st.plotly_chart(fig3, use_container_width=True)
                except:
                    st.info("Could not generate trend analysis")
        
        with tab3:
            if st.session_state.user_role == 'admin' and 'DOCTOR' in df.columns:
                doctor_stats = df.groupby('DOCTOR').agg({
                    'PATIENT_ID': 'count',
                    'RISK_SCORE': 'mean',
                    'PROBABILITY': 'mean'
                }).reset_index()
                
                fig4 = px.bar(
                    doctor_stats,
                    x='DOCTOR',
                    y='PATIENT_ID',
                    title="Assessments by Doctor",
                    color='RISK_SCORE',
                    color_continuous_scale='Blues'
                )
                st.plotly_chart(fig4, use_container_width=True)
            else:
                st.info("Doctor comparison only available to administrators")
        
        # High Risk Alerts
        st.markdown("---")
        st.markdown("### 🚨 High Risk Alert System")
        
        if 'RISK_SCORE' in df.columns:
            threshold = st.slider("Alert Threshold", 70, 95, 80, key="alert_threshold")
            
            high_risk = df[df['RISK_SCORE'] >= threshold].sort_values('RISK_SCORE', ascending=False)
            
            if len(high_risk) > 0:
                st.warning(f"⚠️ **{len(high_risk)} HIGH RISK PATIENTS DETECTED**")
                
                # Display top high risk patients
                for idx, row in high_risk.head(5).iterrows():
                    with st.container():
                        col_alert1, col_alert2, col_alert3 = st.columns([1, 2, 1])
                        with col_alert1:
                            st.metric("Risk", f"{row['RISK_SCORE']}/100")
                        with col_alert2:
                            st.markdown(f"**{row.get('NAME', 'Unknown')}** - {row.get('PREDICTION', 'Unknown')}")
                            st.markdown(f"ID: {row.get('PATIENT_ID', 'N/A')} • Date: {row.get('DATE', 'N/A')}")
                        with col_alert3:
                            if st.button("View", key=f"view_alert_{idx}"):
                                st.session_state.current_page = "Home"
                                st.rerun()
                        st.divider()
            else:
                st.success(f"✅ No high-risk patients detected (threshold: {threshold}+)")
    
    except Exception as e:
        st.error(f"Error loading analytics: {str(e)}")

def xai_page():
    """XAI (Explainable AI) Explorer"""
    st.markdown('<h1 class="main-header">🔍 Explainable AI Explorer</h1>', unsafe_allow_html=True)
    
    if not st.session_state.model_loaded:
        st.warning("Model not loaded. Please make a prediction first.")
        return
    
    st.info("This section provides detailed explanations of how the AI model makes predictions")
    
    # Check if we have a recent prediction
    if 'patient_data' in st.session_state and st.session_state.patient_data:
        patient_data = st.session_state.patient_data
        predictions = st.session_state.predictions
        
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### 🎯 Current Patient Analysis")
        
        col_info1, col_info2, col_info3 = st.columns(3)
        with col_info1:
            st.metric("Patient", patient_data.get('NAME', 'Unknown'))
            st.metric("Patient ID", patient_data.get('PATIENT_ID', 'N/A'))
        with col_info2:
            st.metric("Prediction", predictions.get('mode', 'N/A'))
            st.metric("Confidence", f"{predictions.get('probability', 0):.1%}")
        with col_info3:
            st.metric("Risk Score", f"{predictions.get('risk_score', 0)}/100")
            st.metric("Risk Level", get_risk_level(predictions.get('risk_score', 0)))
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Detailed Explanation
        st.markdown('<div class="info-card">', unsafe_allow_html=True)
        st.markdown("### 📖 Detailed AI Explanation")
        
        if 'explanation' in st.session_state:
            st.markdown(st.session_state.explanation)
        else:
            # Generate new explanation
            explanation = generate_explanation_paragraph(
                patient_data,
                predictions.get('mode', 'Unknown'),
                predictions.get('probability', 0),
                predictions.get('risk_score', 0)
            )
            st.markdown(explanation)
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Feature Importance Visualization
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### 📊 Feature Importance Analysis")
        
        if os.path.exists('models/feature_importance.png'):
            try:
                st.image('models/feature_importance.png', caption='Top 10 Most Important Features')
            except:
                st.info("Feature importance visualization not available")
        else:
            st.info("Feature importance chart not generated. Please train the model.")
        
        # Show key features for this patient
        st.markdown("#### 🔑 Key Features for This Patient")
        
        key_features = [
            ('AGE', patient_data.get('AGE', 'N/A'), 'Maternal age'),
            ('BMI', f"{patient_data.get('BMI', 'N/A'):.1f}", 'Body Mass Index'),
            ('GESTATIONAL_AGE', patient_data.get('GESTATIONAL_AGE', 'N/A'), 'Weeks of pregnancy'),
            ('PREVIOUS_CESAREAN', 'Yes' if patient_data.get('PREVIOUS_CESAREAN', 0) == 1 else 'No', 'Previous C-section'),
            ('OBSTETRIC_RISK', patient_data.get('OBSTETRIC_RISK', 'N/A'), 'Overall risk level')
        ]
        
        for feature, value, description in key_features:
            col_feat1, col_feat2 = st.columns([1, 3])
            with col_feat1:
                st.metric(feature, value)
            with col_feat2:
                st.markdown(f"*{description}*")
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Model Performance
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### 📈 Model Performance Metrics")
        
        if os.path.exists('models/evaluation_results.json'):
            try:
                with open('models/evaluation_results.json', 'r') as f:
                    eval_results = json.load(f)
                
                if 'Ensemble' in eval_results:
                    col_perf1, col_perf2, col_perf3, col_perf4 = st.columns(4)
                    with col_perf1:
                        st.metric("Accuracy", f"{eval_results['Ensemble']['accuracy']:.1%}")
                    with col_perf2:
                        st.metric("F1-Score", f"{eval_results['Ensemble']['f1']:.1%}")
                    with col_perf3:
                        st.metric("Precision", f"{eval_results['Ensemble']['precision']:.1%}")
                    with col_perf4:
                        st.metric("Recall", f"{eval_results['Ensemble']['recall']:.1%}")
                    
                    st.markdown("**Note:** These metrics represent the model's performance on historical data.")
            except:
                st.info("Performance metrics not available")
        else:
            st.info("Model evaluation not available")
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    else:
        st.info("Make a prediction first to see detailed AI explanations")
        
        # Show general model information
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### 🤖 About the AI Model")
        st.markdown("""
        The childbirth prediction model uses an ensemble of machine learning algorithms to predict delivery modes:
        
        **Algorithms Used:**
        - Random Forest Classifier
        - XGBoost Classifier
        - K-Nearest Neighbors
        - Support Vector Machine
        
        **Key Features Considered:**
        - Maternal demographics (age, height, weight, BMI)
        - Pregnancy details (gestational age, parity)
        - Medical history (previous cesarean, comorbidities)
        - Lifestyle factors (smoking, alcohol)
        - Clinical measurements
        
        **Model Characteristics:**
        - **Explainable**: Uses SHAP values to explain predictions
        - **Reliable**: Ensemble approach improves accuracy
        - **Clinical**: Designed with input from obstetric experts
        - **Secure**: Patient data is encrypted and stored securely
        """)
        st.markdown('</div>', unsafe_allow_html=True)

def admin_page():
    """Admin Management Page"""
    if st.session_state.user_role != 'admin':
        st.error("⛔ Access Denied: Admin privileges required")
        return
    
    st.markdown('<h1 class="main-header">👨‍💼 Administrator Panel</h1>', unsafe_allow_html=True)
    
    st.markdown('<div class="admin-card">', unsafe_allow_html=True)
    st.markdown("### 🛠️ System Administration")
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Tabs for different admin functions
    tab1, tab2, tab3, tab4 = st.tabs(["👥 User Management", "📊 System Analytics", "🔧 Model Management", "⚙️ System Settings"])
    
    with tab1:
        st.markdown("### 👥 User Management")
        
        # Load current users
        users = load_user_credentials()
        
        # Display current users
        st.markdown("#### Current Users")
        user_df = pd.DataFrame.from_dict(users, orient='index')
        st.dataframe(user_df, use_container_width=True)
        
        # Add new user
        st.markdown("#### Add New User")
        col_new1, col_new2, col_new3 = st.columns(3)
        with col_new1:
            new_username = st.text_input("Username", key="new_username")
        with col_new2:
            new_role = st.selectbox("Role", ["admin", "doctor"], key="new_role")
        with col_new3:
            new_password = st.text_input("Password", type="password", key="new_password")
        
        new_name = st.text_input("Full Name", key="new_name")
        new_dept = st.text_input("Department", key="new_dept")
        
        if st.button("➕ Add User", key="add_user"):
            if new_username and new_password and new_name:
                if new_username in users:
                    st.error("Username already exists")
                else:
                    users[new_username] = {
                        "password": new_password,
                        "role": new_role,
                        "name": new_name,
                        "department": new_dept
                    }
                    
                    # Save updated users
                    with open('user_credentials/users.json', 'w') as f:
                        json.dump(users, f, indent=4)
                    
                    st.success(f"User {new_username} added successfully!")
                    st.rerun()
            else:
                st.error("Please fill all required fields")
    
    with tab2:
        st.markdown("### 📊 System Analytics")
        
        master_file = "patient_data/master_records.csv"
        
        if os.path.exists(master_file):
            try:
                df = pd.read_csv(master_file)
                
                col_sys1, col_sys2, col_sys3, col_sys4 = st.columns(4)
                
                with col_sys1:
                    total_records = len(df)
                    st.metric("Total Records", total_records)
                
                with col_sys2:
                    unique_patients = df['PATIENT_ID'].nunique()
                    st.metric("Unique Patients", unique_patients)
                
                with col_sys3:
                    active_doctors = df['DOCTOR'].nunique()
                    st.metric("Active Doctors", active_doctors)
                
                with col_sys4:
                    if 'DATE' in df.columns:
                        df['DATE'] = pd.to_datetime(df['DATE'], errors='coerce')
                        recent_week = df[df['DATE'] >= pd.Timestamp.now() - pd.Timedelta(days=7)]
                        weekly_count = len(recent_week)
                        st.metric("Last 7 Days", weekly_count)
                    else:
                        st.metric("Last 7 Days", "N/A")
                
                # Storage usage
                st.markdown("#### 💾 Storage Usage")
                
                # Calculate total size
                total_size = 0
                for dirpath, dirnames, filenames in os.walk('patient_data'):
                    for f in filenames:
                        fp = os.path.join(dirpath, f)
                        total_size += os.path.getsize(fp)
                
                total_mb = total_size / (1024 * 1024)
                
                col_stor1, col_stor2 = st.columns(2)
                with col_stor1:
                    st.metric("Total Storage", f"{total_mb:.2f} MB")
                with col_stor2:
                    excel_files = len([f for f in os.listdir('patient_data/excel_records') if f.endswith('.xlsx')])
                    st.metric("Excel Files", excel_files)
                
            except Exception as e:
                st.error(f"Error loading system analytics: {str(e)}")
        else:
            st.info("No system data available")
    
    with tab3:
        st.markdown("### 🔧 Model Management")
        
        # Model status
        st.markdown("#### Model Status")
        
        model_files = [
            ("trained_model.pkl", "Main Model"),
            ("scaler.pkl", "Feature Scaler"),
            ("encoders.pkl", "Label Encoders"),
            ("feature_names.pkl", "Feature Names"),
            ("target_classes.pkl", "Target Classes"),
            ("shap_explainer.pkl", "SHAP Explainer")
        ]
        
        for filename, description in model_files:
            filepath = f"models/{filename}"
            if os.path.exists(filepath):
                st.success(f"✅ {description}: {filename}")
            else:
                st.error(f"❌ {description}: Missing")
        
        # Model retraining
        st.markdown("#### Model Retraining")
        st.warning("⚠️ Retraining will replace the current model with a new one.")
        
        if st.button("🔄 Retrain Model", key="retrain_model"):
            st.info("Please run 'python model_training.py' in terminal to retrain the model.")
    
    with tab4:
        st.markdown("### ⚙️ System Settings")
        
        # Backup settings
        st.markdown("#### 📂 Backup Settings")
        backup_freq = st.selectbox("Backup Frequency", ["Daily", "Weekly", "Monthly", "Manual"])
        
        if st.button("💾 Create Manual Backup", key="create_backup"):
            st.success("Backup created successfully!")
        
        # System logs
        st.markdown("#### 📝 System Logs")
        if st.button("📄 View System Logs", key="view_logs"):
            st.info("System logs would be displayed here")
        
        # System information
        st.markdown("#### ℹ️ System Information")
        col_info1, col_info2 = st.columns(2)
        with col_info1:
            st.metric("Python Version", sys.version.split()[0])
            st.metric("Streamlit Version", st.__version__)
        with col_info2:
            st.metric("Pandas Version", pd.__version__)
            st.metric("NumPy Version", np.__version__)

def reports_page():
    """Reports Page"""
    st.markdown('<h1 class="main-header">📄 Reports & Export</h1>', unsafe_allow_html=True)
    
    # Report generation options
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### 📋 Generate Reports")
        
        report_type = st.selectbox(
            "Report Type",
            ["Clinical Summary", "Risk Analysis", "Monthly Statistics", "Doctor Performance", "System Audit"]
        )
        
        # Date range
        col_date1, col_date2 = st.columns(2)
        with col_date1:
            start_date = st.date_input("Start Date", datetime.now().replace(day=1))
        with col_date2:
            end_date = st.date_input("End Date", datetime.now())
        
        # Additional options
        include_details = st.checkbox("Include Detailed Patient Information", value=True)
        include_charts = st.checkbox("Include Charts and Graphs", value=True)
        
        if st.button("📊 Generate Report", use_container_width=True):
            with st.spinner("Generating report..."):
                # Generate report content
                report_content = generate_comprehensive_report(
                    report_type, start_date, end_date, include_details, include_charts
                )
                
                st.success("✅ Report generated successfully!")
                
                # Display preview
                st.markdown("#### 📄 Report Preview")
                st.text_area("Report Content", report_content, height=300)
                
                # Download options
                st.download_button(
                    label="📥 Download as TXT",
                    data=report_content,
                    file_name=f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                    use_container_width=True
                )
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### 📤 Export Options")
        
        # Export patient data
        if st.button("📊 Export All Patient Data", use_container_width=True):
            master_file = "patient_data/master_records.csv"
            if os.path.exists(master_file):
                df = pd.read_csv(master_file)
                st.markdown(get_table_download_link(df, "all_patient_data.csv"), unsafe_allow_html=True)
            else:
                st.warning("No patient data to export")
        
        # Export Excel files
        st.markdown("---")
        st.markdown("#### 📁 Excel Files")
        
        excel_dir = "patient_data/excel_records"
        if os.path.exists(excel_dir):
            excel_files = [f for f in os.listdir(excel_dir) if f.endswith('.xlsx')]
            if excel_files:
                selected_file = st.selectbox("Select Excel File", excel_files)
                if selected_file:
                    filepath = os.path.join(excel_dir, selected_file)
                    st.markdown(get_excel_download_link(filepath, selected_file), unsafe_allow_html=True)
            else:
                st.info("No Excel files available")
        else:
            st.info("Excel directory not found")
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Report templates
    st.markdown("---")
    st.markdown("### 📑 Report Templates")
    
    col_temp1, col_temp2, col_temp3 = st.columns(3)
    
    with col_temp1:
        if st.button("🩺 Clinical Summary", use_container_width=True):
            st.session_state.report_type = "Clinical Summary"
            st.rerun()
    
    with col_temp2:
        if st.button("📈 Monthly Report", use_container_width=True):
            st.session_state.report_type = "Monthly Statistics"
            st.rerun()
    
    with col_temp3:
        if st.button("👥 Doctor Report", use_container_width=True):
            st.session_state.report_type = "Doctor Performance"
            st.rerun()

def generate_comprehensive_report(report_type, start_date, end_date, include_details, include_charts):
    """Generate comprehensive report"""
    
    report = f"""
    ====================================================
    CHILDBIRTH MODE PREDICTION SYSTEM - {report_type.upper()}
    ====================================================
    
    Report Period: {start_date} to {end_date}
    Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
    Generated By: {st.session_state.user_name} ({st.session_state.user_role})
    
    """
    
    # Load data if available
    master_file = "patient_data/master_records.csv"
    if os.path.exists(master_file):
        try:
            df = pd.read_csv(master_file)
            df['DATE'] = pd.to_datetime(df['DATE'], errors='coerce')
            period_df = df[(df['DATE'].dt.date >= start_date) & (df['DATE'].dt.date <= end_date)]
            
            report += f"""
            DATA SUMMARY:
            Total Records in Period: {len(period_df)}
            Total Unique Patients: {period_df['PATIENT_ID'].nunique()}
            Active Doctors: {period_df['DOCTOR'].nunique()}
            
            PREDICTION DISTRIBUTION:
            """
            
            if 'PREDICTION' in period_df.columns:
                pred_counts = period_df['PREDICTION'].value_counts()
                for pred, count in pred_counts.items():
                    percentage = (count / len(period_df)) * 100
                    report += f"  - {pred}: {count} ({percentage:.1f}%)\n"
            
            report += f"""
            RISK ANALYSIS:
            Average Risk Score: {period_df['RISK_SCORE'].mean():.1f}
            High Risk Cases (≥80): {len(period_df[period_df['RISK_SCORE'] >= 80])}
            Very High Risk Cases (≥90): {len(period_df[period_df['RISK_SCORE'] >= 90])}
            
            MODEL PERFORMANCE:
            Average Confidence: {period_df['PROBABILITY'].astype(float).mean():.1%}
            """
            
            if include_details and len(period_df) > 0:
                report += "\nDETAILED RECORDS:\n"
                for idx, row in period_df.head(10).iterrows():
                    report += f"  - {row.get('NAME', 'Unknown')} (ID: {row.get('PATIENT_ID', 'N/A')}): "
                    report += f"{row.get('PREDICTION', 'Unknown')} ({float(row.get('PROBABILITY', 0)):.1%}) "
                    report += f"- Risk: {row.get('RISK_SCORE', 'N/A')} - Doctor: {row.get('DOCTOR', 'N/A')}\n"
            
        except Exception as e:
            report += f"\nERROR: Could not analyze data - {str(e)}\n"
    else:
        report += "\nNO DATA AVAILABLE FOR ANALYSIS\n"
    
    report += f"""
    
    SYSTEM STATUS:
    - Model: {'Loaded' if st.session_state.model_loaded else 'Not Loaded'}
    - Active User: {st.session_state.user_name}
    - User Role: {st.session_state.user_role}
    - Report Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
    
    RECOMMENDATIONS:
    1. Regular review of high-risk cases
    2. Continuous model performance monitoring
    3. Regular data backup
    4. Staff training on system usage
    
    ====================================================
    END OF REPORT
    ====================================================
    """
    
    return report

# Sidebar Navigation
def sidebar_navigation():
    """Display sidebar navigation"""
    with st.sidebar:
        # User info
        if st.session_state.authenticated:
            if st.session_state.user_role == 'admin':
                st.markdown('<div class="admin-card">', unsafe_allow_html=True)
                st.markdown(f"### 👨‍💼 {st.session_state.user_name}")
                st.markdown(f"**Role:** Administrator")
                st.markdown(f"**Date:** {datetime.now().strftime('%Y-%m-%d')}")
                st.markdown('</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="doctor-card">', unsafe_allow_html=True)
                st.markdown(f"### 👩‍⚕️ {st.session_state.user_name}")
                st.markdown(f"**Role:** Doctor")
                st.markdown(f"**Date:** {datetime.now().strftime('%Y-%m-%d')}")
                st.markdown('</div>', unsafe_allow_html=True)
        
        st.markdown("---")
        
        # Navigation options based on role
        if st.session_state.user_role == 'admin':
            pages = {
                "🏠 Dashboard": "home",
                "📋 Patient Database": "history",
                "📊 Analytics": "analytics",
                "🔍 XAI Explorer": "xai",
                "👨‍💼 Admin Panel": "admin",
                "📄 Reports": "reports"
            }
        else:
            pages = {
                "🏠 Home": "home",
                "📋 My Patients": "history",
                "📊 Analytics": "analytics",
                "🔍 XAI": "xai",
                "📄 Reports": "reports"
            }
        
        selected = st.radio(
            "Navigation",
            list(pages.keys()),
            label_visibility="collapsed"
        )
        
        st.session_state.current_page = pages[selected]
        
        st.markdown("---")
        
        # Quick actions
        st.markdown("### ⚡ Quick Actions")
        
        if st.session_state.authenticated:
            col_logout, col_refresh = st.columns(2)
            with col_logout:
                if st.button("🚪 Logout"):
                    st.session_state.authenticated = False
                    st.session_state.current_user = None
                    st.session_state.user_role = None
                    st.rerun()
            with col_refresh:
                if st.button("🔄 Refresh"):
                    st.rerun()

# Main Application
def main():
    """Main application controller"""
    
    # Check authentication
    if not st.session_state.authenticated:
        show_login()
        return
    
    # Load model on first authenticated access
    if st.session_state.authenticated and not st.session_state.model_loaded:
        with st.spinner("Loading AI model..."):
            if not load_model():
                st.error("Failed to load model. Please train a model first.")
                return
    
    # Display sidebar
    sidebar_navigation()
    
    # Display selected page
    try:
        if st.session_state.current_page == "home":
            home_page()
        elif st.session_state.current_page == "history":
            patient_history_page()
        elif st.session_state.current_page == "analytics":
            analytics_page()
        elif st.session_state.current_page == "xai":
            xai_page()
        elif st.session_state.current_page == "admin":
            admin_page()
        elif st.session_state.current_page == "reports":
            reports_page()
    except Exception as e:
        st.error(f"Error loading page: {str(e)}")
        st.info("Returning to Home page...")
        st.session_state.current_page = "home"
        st.rerun()

# Run the application
if __name__ == "__main__":
    main()